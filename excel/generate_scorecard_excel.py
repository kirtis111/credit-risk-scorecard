"""
generate_scorecard_excel.py
────────────────────────────
Exports the credit risk scorecard to a professionally formatted Excel workbook.

Sheets:
  1. Scorecard Points  — Bin-level point allocation table
  2. WoE Summary       — IV table with strength ratings
  3. Score Distribution — Performance by score band
  4. Champion vs Challenger — Model comparison
  5. IFRS9 Provisions  — ECL staging summary
  6. Monitoring Template — PSI/CSI tracking template
"""

import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill, numbers)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import date
import warnings
warnings.filterwarnings("ignore")


# ── Style constants ────────────────────────────────────────────
NAVY_HEX   = "003366"
BLUE_HEX   = "2563EB"
LTBLUE_HEX = "DBEAFE"
GREEN_HEX  = "16A34A"
LTGRN_HEX  = "DCFCE7"
AMBER_HEX  = "D97706"
LTAMB_HEX  = "FEF3C7"
RED_HEX    = "DC2626"
LTRED_HEX  = "FEE2E2"
GREY_HEX   = "F1F5F9"
DARKGRY    = "475569"
WHITE_HEX  = "FFFFFF"

def navy_fill():  return PatternFill("solid", fgColor=NAVY_HEX)
def blue_fill():  return PatternFill("solid", fgColor=BLUE_HEX)
def ltblue_fill():return PatternFill("solid", fgColor=LTBLUE_HEX)
def green_fill(): return PatternFill("solid", fgColor=GREEN_HEX)
def ltgrn_fill(): return PatternFill("solid", fgColor=LTGRN_HEX)
def amber_fill(): return PatternFill("solid", fgColor=AMBER_HEX)
def red_fill():   return PatternFill("solid", fgColor=RED_HEX)
def grey_fill():  return PatternFill("solid", fgColor=GREY_HEX)

def header_font(size=10, color=WHITE_HEX, bold=True):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=9, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def center_align(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left_align():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def thin_border():
    side = Side(style="thin", color="CBD5E1")
    return Border(left=side, right=side, top=side, bottom=side)

def thick_border():
    thick = Side(style="medium", color=NAVY_HEX)
    thin  = Side(style="thin",   color="CBD5E1")
    return Border(left=thick, right=thick, top=thick, bottom=thick)

def rag_fill(status):
    return green_fill() if status=="Green" else (amber_fill() if status=="Amber" else red_fill())

def rag_font(status):
    return Font(name="Calibri", size=9, bold=True, color=WHITE_HEX)

def style_header_row(ws, row, col_start, col_end, fill=None, font=None, height=22):
    fill = fill or navy_fill()
    font = font or header_font()
    ws.row_dimensions[row].height = height
    for col in range(col_start, col_end+1):
        cell = ws.cell(row=row, column=col)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = center_align()
        cell.border    = thin_border()

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ── Sheet 1: Scorecard Points ──────────────────────────────────
def write_scorecard_sheet(wb: Workbook, scorecard_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Scorecard Points")
    ws.sheet_properties.tabColor = NAVY_HEX

    # Title block
    ws.merge_cells("A1:H1")
    ws["A1"] = "CREDIT RISK SCORECARD — POINT ALLOCATION TABLE"
    ws["A1"].fill  = navy_fill()
    ws["A1"].font  = header_font(13)
    ws["A1"].alignment = center_align()
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Model Version: v1.0  |  Scale: 300–850  |  Base Score: 600  |  PDO: 20  |  Generated: {date.today()}"
    ws["A2"].fill  = blue_fill()
    ws["A2"].font  = header_font(9)
    ws["A2"].alignment = center_align()
    ws.row_dimensions[2].height = 16

    # Column headers
    headers = ["Feature", "Bin Description", "N Observations", "Default Rate %",
               "WoE Value", "LR Coefficient", "Score Points", "Running Total"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=3, column=j, value=h)
    style_header_row(ws, 3, 1, 8)

    # Data
    row_n = 4
    prev_feature = None
    running_total_per_feat = {}
    for _, r in scorecard_df.iterrows():
        feat = r["Feature"]
        if feat != prev_feature:
            # Feature separator row
            ws.merge_cells(f"A{row_n}:H{row_n}")
            cell = ws.cell(row=row_n, column=1,
                           value=f"▶  {feat.replace('_WoE','').upper()}")
            cell.fill = ltblue_fill()
            cell.font = Font(name="Calibri", size=9, bold=True, color=NAVY_HEX)
            cell.alignment = left_align()
            ws.row_dimensions[row_n].height = 16
            row_n += 1
            prev_feature = feat

        is_even = row_n % 2 == 0
        fill_row = grey_fill() if is_even else PatternFill("solid", fgColor=WHITE_HEX)

        vals = [feat.replace("_WoE",""), r["Bin"], int(r["N"]),
                round(r["EventRate"], 2), round(r["WoE"], 4),
                round(r["Coefficient"], 6), round(r["Points"], 1), ""]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row_n, column=j, value=v)
            cell.fill   = fill_row
            cell.font   = body_font()
            cell.border = thin_border()
            cell.alignment = center_align() if j > 1 else left_align()
            if j in (5, 6, 7):
                cell.number_format = "0.0000" if j < 7 else "0.0"

        # Colour code WoE cell
        woe_cell = ws.cell(row=row_n, column=5)
        woe_cell.fill = ltgrn_fill() if r["WoE"] > 0 else PatternFill("solid", fgColor=LTRED_HEX)

        row_n += 1

    for col, width in [("A",20),("B",35),("C",16),("D",15),("E",12),("F",15),("G",13),("H",13)]:
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"


# ── Sheet 2: WoE IV Summary ────────────────────────────────────
def write_iv_sheet(wb: Workbook, iv_table: pd.DataFrame) -> None:
    ws = wb.create_sheet("WoE IV Summary")
    ws.sheet_properties.tabColor = "2563EB"

    ws.merge_cells("A1:F1")
    ws["A1"] = "INFORMATION VALUE (IV) — FEATURE PREDICTIVE POWER"
    ws["A1"].fill = navy_fill(); ws["A1"].font = header_font(12)
    ws["A1"].alignment = center_align(); ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = "IV < 0.02: Useless  |  0.02–0.10: Weak  |  0.10–0.30: Medium  |  0.30–0.50: Strong  |  > 0.50: Suspicious"
    ws["A2"].fill = ltblue_fill(); ws["A2"].font = body_font(8)
    ws["A2"].alignment = center_align()

    headers = ["Rank","Feature","Information Value","Predictive Strength","# Bins","Selected"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=3, column=j, value=h)
    style_header_row(ws, 3, 1, 6)

    iv_sorted = iv_table.sort_values("IV", ascending=False).reset_index(drop=True)
    strength_fills = {
        "Useless":   PatternFill("solid", fgColor="E2E8F0"),
        "Weak":      PatternFill("solid", fgColor=LTAMB_HEX),
        "Medium":    PatternFill("solid", fgColor=LTBLUE_HEX),
        "Strong":    PatternFill("solid", fgColor=LTGRN_HEX),
        "Suspicious (check leakage)": PatternFill("solid", fgColor=LTRED_HEX),
    }
    iv_minimum = 0.02
    for i, r in iv_sorted.iterrows():
        row_n = i + 4
        selected = "✓ Selected" if r["IV"] >= iv_minimum else "✗ Excluded"
        vals = [i+1, r["Feature"].replace("_WoE",""), r["IV"], r["Strength"],
                r.get("N_Bins", "—"), selected]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row_n, column=j, value=v)
            cell.fill = (grey_fill() if i % 2 == 0 else PatternFill("solid", fgColor=WHITE_HEX))
            cell.font = body_font(bold=(j==2))
            cell.border = thin_border()
            cell.alignment = center_align() if j != 2 else left_align()
        # Colour Strength cell
        strength_cell = ws.cell(row=row_n, column=4)
        strength_cell.fill = strength_fills.get(r["Strength"], grey_fill())
        # Colour Selected cell
        sel_cell = ws.cell(row=row_n, column=6)
        sel_cell.fill = ltgrn_fill() if "✓" in selected else PatternFill("solid", fgColor=LTRED_HEX)
        sel_cell.font = Font(name="Calibri", size=9, bold=True,
                              color=GREEN_HEX if "✓" in selected else RED_HEX)

    for col, width in [("A",7),("B",28),("C",18),("D",30),("E",10),("F",14)]:
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"

    # Add bar chart for IV
    chart = BarChart()
    chart.type = "bar"; chart.title = "Information Value by Feature"
    chart.y_axis.title = "Feature"; chart.x_axis.title = "IV"
    data_ref = Reference(ws, min_col=3, min_row=3, max_row=3+len(iv_sorted))
    cats_ref = Reference(ws, min_col=2, min_row=4, max_row=3+len(iv_sorted))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4; chart.style = 10
    chart.width = 16; chart.height = 12
    ws.add_chart(chart, f"H3")


# ── Sheet 3: Score Distribution ────────────────────────────────
def write_score_dist_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Score Distribution")
    ws.sheet_properties.tabColor = GREEN_HEX

    ws.merge_cells("A1:I1")
    ws["A1"] = "SCORE DISTRIBUTION ANALYSIS — CREDIT DECISION BANDS"
    ws["A1"].fill = navy_fill(); ws["A1"].font = header_font(12)
    ws["A1"].alignment = center_align(); ws.row_dimensions[1].height = 28

    bands = [
        ("300–549", "Very High Risk",  "DECLINE",             RED_HEX,   LTRED_HEX,   12.0, 65.0),
        ("550–579", "High Risk",       "DECLINE",             RED_HEX,   LTRED_HEX,   8.0,  48.0),
        ("580–619", "Medium Risk",     "REFER TO CREDIT",     AMBER_HEX, LTAMB_HEX,   5.2,  32.0),
        ("620–659", "Acceptable Risk", "CONDITIONAL APPROVE", "CA8A04",  "FEF9C3",    3.1,  18.5),
        ("660–699", "Low Risk",        "APPROVE",             GREEN_HEX, LTGRN_HEX,   1.8,  10.2),
        ("700–749", "Very Low Risk",   "APPROVE",             GREEN_HEX, LTGRN_HEX,   0.9,  6.5),
        ("750–850", "Minimal Risk",    "APPROVE (FAST TRACK)",GREEN_HEX, LTGRN_HEX,   0.4,  3.1),
    ]
    headers = ["Score Band","Risk Category","Decision","% Population","Cum % Pop",
               "Default Rate %","Est. PD","Avg ECL Rate %"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=2, column=j, value=h)
    style_header_row(ws, 2, 1, 8)

    # Synthetic pop distribution
    pop_pcts = [3.2, 5.8, 12.4, 21.6, 28.5, 18.3, 10.2]
    cum_pop  = 0.0
    for i, (band, cat, dec, fc, lc, dr, pd_est) in enumerate(bands):
        row_n = i + 3
        cum_pop += pop_pcts[i]
        ecl_rate = dr * 0.35
        vals = [band, cat, dec, f"{pop_pcts[i]:.1f}%", f"{cum_pop:.1f}%",
                f"{dr:.1f}%", f"{pd_est:.1f}%", f"{ecl_rate:.2f}%"]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row_n, column=j, value=v)
            cell.fill = PatternFill("solid", fgColor=lc)
            cell.font = Font(name="Calibri", size=9,
                             bold=(j in [1,3]), color=fc if j in [1,2,3] else "000000")
            cell.border = thin_border()
            cell.alignment = center_align()
        ws.row_dimensions[row_n].height = 18

    # OSFI threshold annotations
    ws.merge_cells("A11:H11")
    ws["A11"] = ("⚠️  OSFI B-20 Decision Rules: Score < 580 = Decline  |  "
                  "580–619 = Refer to Credit Officer  |  620–659 = Conditional  |  ≥ 660 = Approve")
    ws["A11"].fill = ltblue_fill()
    ws["A11"].font = Font(name="Calibri", size=8, bold=True, color=NAVY_HEX)
    ws["A11"].alignment = center_align()

    for col, width in [("A",12),("B",20),("C",22),("D",14),("E",12),
                        ("F",15),("G",12),("H",15)]:
        ws.column_dimensions[col].width = width


# ── Sheet 4: Champion-Challenger ───────────────────────────────
def write_comparison_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Champion vs Challenger")
    ws.sheet_properties.tabColor = "7C3AED"

    ws.merge_cells("A1:F1")
    ws["A1"] = "CHAMPION vs CHALLENGER — MODEL COMPARISON (OSFI E-23)"
    ws["A1"].fill = navy_fill(); ws["A1"].font = header_font(12)
    ws["A1"].alignment = center_align(); ws.row_dimensions[1].height = 28

    headers = ["Metric","Champion (LR Scorecard)","Challenger (XGBoost)","Difference","Verdict","Threshold"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=2, column=j, value=h)
    style_header_row(ws, 2, 1, 6)

    # Illustrative results (replace with actual computed values)
    rows = [
        ("AUC-ROC",      0.8312, 0.8687, ">0.75",  True),
        ("Gini",         0.6624, 0.7374, ">0.50",  True),
        ("KS Statistic", 0.5218, 0.5834, ">0.40",  True),
        ("Brier Score",  0.1423, 0.1218, "<0.25",  False),
        ("Sensitivity",  0.7234, 0.7612, ">0.65",  True),
        ("Specificity",  0.8156, 0.8398, ">0.75",  True),
        ("F1 Score",     0.6812, 0.7124, ">0.60",  True),
    ]
    for i, (metric, champ, chall, thresh, higher_better) in enumerate(rows):
        row_n = i + 3
        diff = chall - champ
        if metric == "Brier Score":
            verdict = "✓ Challenger Superior" if diff < -0.01 else ("= Comparable" if abs(diff) <= 0.01 else "✓ Champion Superior")
        else:
            verdict = "✓ Challenger Superior" if diff > 0.01 else ("= Comparable" if abs(diff) <= 0.01 else "✓ Champion Superior")

        fill_v = ltgrn_fill() if "Challenger" in verdict else (ltblue_fill() if "Comparable" in verdict else PatternFill("solid", fgColor=LTAMB_HEX))
        row_fill = grey_fill() if i % 2 == 0 else PatternFill("solid", fgColor=WHITE_HEX)

        for j, v in enumerate([metric, champ, chall, round(diff,4), verdict, thresh], 1):
            cell = ws.cell(row=row_n, column=j, value=v)
            cell.fill   = fill_v if j in [4,5] else row_fill
            cell.font   = body_font(bold=(j in [1,5]))
            cell.border = thin_border()
            cell.alignment = center_align() if j > 1 else left_align()
            if j in [2,3,4]:
                cell.number_format = "0.0000"

    for col, width in [("A",20),("B",22),("C",22),("D",14),("E",26),("F",12)]:
        ws.column_dimensions[col].width = width


# ── Sheet 5: IFRS 9 Provisions ─────────────────────────────────
def write_ifrs9_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("IFRS9 Provisions")
    ws.sheet_properties.tabColor = "DC2626"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"IFRS 9 ECL PROVISION REGISTER — {date.today().strftime('%B %Y').upper()}"
    ws["A1"].fill = navy_fill(); ws["A1"].font = header_font(12)
    ws["A1"].alignment = center_align(); ws.row_dimensions[1].height = 28

    headers = ["IFRS 9 Stage","Description","N Exposures","EAD (CAD)",
               "ECL — Base","ECL — FLI Adj","Coverage %","RAG"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=2, column=j, value=h)
    style_header_row(ws, 2, 1, 8)

    stages = [
        (1,"Stage 1 — Performing (12M ECL)",  3812, 18_420_000, 221_040, 198_936, ltgrn_fill(), GREEN_HEX),
        (2,"Stage 2 — Underperforming (Lifetime)", 912, 4_380_000, 374_580, 411_750, PatternFill("solid", fgColor=LTAMB_HEX), AMBER_HEX),
        (3,"Stage 3 — Non-Performing (Lifetime)", 236, 1_800_000, 504_000, 554_400, PatternFill("solid", fgColor=LTRED_HEX), RED_HEX),
    ]
    for i, (stage, desc, n, ead, ecl_base, ecl_fli, fill_color, fc) in enumerate(stages):
        row_n = i + 3
        cov = ecl_base / ead * 100
        rag = "Green" if stage == 1 else ("Amber" if stage == 2 else "Red")
        for j, v in enumerate([stage, desc, n, ead, ecl_base, ecl_fli, round(cov,2), f"● {rag}"], 1):
            cell = ws.cell(row=row_n, column=j, value=v)
            cell.fill   = fill_color
            cell.font   = Font(name="Calibri", size=9, bold=(j in [1,2,8]),
                                color=fc if j in [1,2,8] else "000000")
            cell.border = thin_border()
            cell.alignment = center_align() if j != 2 else left_align()
            if j in [4,5,6]:
                cell.number_format = '#,##0'
            if j == 7:
                cell.number_format = '0.00"%"'
        ws.row_dimensions[row_n].height = 20

    # Totals row
    row_n = 6
    total_ead = sum(s[3] for s in stages)
    total_ecl = sum(s[4] for s in stages)
    total_fli = sum(s[5] for s in stages)
    total_cov = total_ecl / total_ead * 100
    for j, v in enumerate(["TOTAL","Portfolio Total",
                             sum(s[2] for s in stages),
                             total_ead, total_ecl, total_fli, round(total_cov,2),"—"], 1):
        cell = ws.cell(row=row_n, column=j, value=v)
        cell.fill = navy_fill(); cell.font = header_font(9)
        cell.border = thin_border(); cell.alignment = center_align()
        if j in [4,5,6]: cell.number_format = '#,##0'
        if j == 7: cell.number_format = '0.00"%"'

    for col, width in [("A",10),("B",35),("C",15),("D",18),
                        ("E",18),("F",18),("G",13),("H",12)]:
        ws.column_dimensions[col].width = width


# ── Sheet 6: Monitoring Template ──────────────────────────────
def write_monitoring_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("PSI Monitoring")
    ws.sheet_properties.tabColor = "0EA5E9"

    ws.merge_cells("A1:J1")
    ws["A1"] = "MONTHLY PSI/CSI MODEL MONITORING TEMPLATE — OSFI E-23"
    ws["A1"].fill = navy_fill(); ws["A1"].font = header_font(12)
    ws["A1"].alignment = center_align(); ws.row_dimensions[1].height = 28

    headers = ["Period","N Population","Mean Score","Std Score","Default Rate %",
               "Expected PD %","AUC-ROC","PSI Score","CSI (Avg)","RAG Status"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=2, column=j, value=h)
    style_header_row(ws, 2, 1, 10)

    # 12 months of illustrative monitoring data
    months = pd.date_range("2024-01-01", periods=12, freq="MS")
    base_score = 638; base_psi = 0.005
    for i, mo in enumerate(months):
        row_n = i + 3
        drift  = i * 1.5
        score  = base_score - drift
        psi_v  = base_psi + i * 0.011 + np.random.uniform(0, 0.007)
        dr     = 0.198 + i * 0.003
        exp_pd = 0.195 + i * 0.002
        auc    = 0.832 - i * 0.003
        rag    = "🟢 Green" if psi_v < 0.10 else ("🟡 Amber" if psi_v < 0.25 else "🔴 Red")
        fill_r = grey_fill() if i % 2 == 0 else PatternFill("solid", fgColor=WHITE_HEX)
        vals   = [mo.strftime("%b %Y"), 450+i*5, round(score,1), round(65+i*0.5,1),
                  round(dr*100,2), round(exp_pd*100,2), round(auc,4),
                  round(psi_v,4), round(psi_v*0.8,4), rag]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row_n, column=j, value=v)
            cell.fill   = fill_r
            cell.font   = body_font()
            cell.border = thin_border()
            cell.alignment = center_align()
        rag_cell = ws.cell(row=row_n, column=10)
        rag_cell.fill = (ltgrn_fill() if "Green" in rag else
                         (PatternFill("solid",fgColor=LTAMB_HEX) if "Amber" in rag else
                          PatternFill("solid",fgColor=LTRED_HEX)))
        rag_cell.font = Font(name="Calibri", size=9, bold=True,
                              color=(GREEN_HEX if "Green" in rag else
                                     (AMBER_HEX if "Amber" in rag else RED_HEX)))

    # Thresholds note
    row_n = 16
    ws.merge_cells(f"A{row_n}:J{row_n}")
    ws[f"A{row_n}"] = ("PSI Thresholds: < 0.10 = Green (No Action)  |  "
                        "0.10–0.25 = Amber (Increase Monitoring)  |  "
                        "> 0.25 = Red (Rebuild Required — Escalate to CRO)")
    ws[f"A{row_n}"].fill = ltblue_fill()
    ws[f"A{row_n}"].font = Font(name="Calibri", size=8, bold=True, color=NAVY_HEX)
    ws[f"A{row_n}"].alignment = center_align()

    for col, width in [("A",12),("B",14),("C",13),("D",12),("E",16),
                        ("F",16),("G",12),("H",12),("I",12),("J",14)]:
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"


# ── Main builder ──────────────────────────────────────────────
def build_excel_workbook(output_path: str = "excel/Scorecard_v1.0.xlsx",
                          scorecard_df: pd.DataFrame = None,
                          iv_table: pd.DataFrame = None) -> str:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # Generate synthetic data if not provided
    if scorecard_df is None:
        np.random.seed(42)
        features   = ["DEBTINC","DELINQ","DEROG","CLAGE","LOAN","VALUE","YOJ","NINQ","CLNO"]
        rows = []
        for f in features:
            for b in range(5):
                woe = np.random.uniform(-1.2, 1.2)
                pts = -woe * 28.85 + 600/9
                rows.append({"Feature": f"{f}_WoE", "Bin": f"Bin {b+1}", "N": np.random.randint(200,800),
                              "EventRate": np.random.uniform(0.05, 0.45),
                              "WoE": round(woe,4), "Coefficient": round(np.random.uniform(-0.5,0.5),6),
                              "Points": round(pts,1)})
        scorecard_df = pd.DataFrame(rows)

    if iv_table is None:
        feats  = ["DEBTINC","DELINQ","DEROG","CLAGE","LOAN","VALUE","YOJ","NINQ","CLNO","JOB","REASON"]
        iv_vals= [0.482, 0.341, 0.298, 0.201, 0.153, 0.148, 0.112, 0.089, 0.061, 0.034, 0.019]
        strs   = ["Strong","Strong","Strong","Medium","Medium","Medium","Medium","Weak","Weak","Weak","Useless"]
        iv_table = pd.DataFrame({"Feature":feats,"IV":iv_vals,"Strength":strs,"N_Bins":[9,5,4,9,9,8,7,6,6,4,3]})

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    write_scorecard_sheet(wb, scorecard_df)
    write_iv_sheet(wb, iv_table)
    write_score_dist_sheet(wb)
    write_comparison_sheet(wb)
    write_ifrs9_sheet(wb)
    write_monitoring_sheet(wb)

    wb.save(output_path)
    print(f" Excel workbook saved: {output_path}")
    print(f"   Sheets: {[s.title for s in wb.worksheets]}")
    return output_path


# ──────────────────────────────────────────────────────────────
# generate_scorecard_excel.py is a LIBRARY MODULE.
# Excel generation has a single entry point: notebooks/06_Model_Validation_Report.py
# Do NOT run this file directly — call build_excel_workbook() from notebook 06.
# ──────────────────────────────────────────────────────────────
